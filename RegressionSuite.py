import os
import sys
from glob import glob
import pandas as pd
from openpyxl import load_workbook
from time import time


def check_same_cols(df1, df2):
    return all(df1.columns == df2.columns)

def compare_two_dfs(df_ideal, df_transformed, columns_to_ignore=[]):
    if_same_cols = check_same_cols(df_ideal, df_transformed)
    df_ideal = df_ideal.fillna(0)
    df_transformed = df_transformed.fillna(0)
    
    if if_same_cols:
        col_set = set(df_ideal.columns) - set(columns_to_ignore)

        for col in col_set:
            temp = df_ideal[col] != df_transformed[col]

            if sum(temp) != 0:
                return (
                    False,
                    f"Mismatch in Column {col}\n{df_ideal[temp][col]} is not equal to \n{df_transformed[temp][col]}",
                )

        return True, ""


def get_comparable_sheetnames(workbook):
    return [
        sheetname
        for sheetname in workbook.sheetnames
        if not (
            sheetname.startswith("$")
            or sheetname.startswith("~")
            or sheetname.startswith("#")
        )
    ]


def if_same_sheets(wb1, wb2):
    wb1_sheets = get_comparable_sheetnames(wb1)
    wb2_sheets = get_comparable_sheetnames(wb2)
    return set(wb1_sheets) == set(wb2_sheets), wb1_sheets


def if_same_color_sheets(wb1, wb2):
    n = len(wb1.worksheets)   
    sheetnames = [name for name in wb2.sheetnames]
    sheetcolors = [sheet.sheet_properties.tabColor for sheet in wb2.worksheets]
    msg = ""
    
    for i in range(0, n):
        if(wb1.worksheets[i].sheet_properties.tabColor != wb2.worksheets[i].sheet_properties.tabColor):
            msg += f"Sheet colors do not match in {wb1.worksheets[i]} and {wb2.worksheets[i]}"
            
        if((sheetnames[i].startswith('~') or sheetnames[i].startswith('#')) and not (sheetcolors[i] == None or sheetcolors[i].rgb == 'FFFFFFFF')):
            msg += f"\n{sheetnames[i]} does not have the appropriate color coding"
        elif(sheetnames[i].startswith('$') and not (sheetcolors[i].rgb == 'FFFF0000')):
            msg += f"\n{sheetnames[i]} does not have the appropriate color coding"
        elif(sheetnames[i].startswith(r'\W') and not (sheetcolors[i].rgb == 'FFFFFF00')):
            msg += f"\n{sheetnames[i]} does not have the appropriate color coding"
    
    return (True, msg) if msg == "" else (False, msg)


def compare_excel_files(transformed_file, ideal_folder):
    cur_sheet = None
    try:
        exists, ideal_file = os.path.exists(f"{ideal_folder}/expected_{os.path.basename(transformed_file)}"), f"{ideal_folder}/expected_{os.path.basename(transformed_file)}"
        if(exists):
            wb1 = load_workbook(ideal_file)
            wb2 = load_workbook(transformed_file)

            same_sheets, sheets = if_same_sheets(wb1, wb2)
            if same_sheets:
                sheet_color, sheet_color_message = if_same_color_sheets(wb1, wb2)
                for sheet in sheets:
                    cur_sheet = sheet
                    df_ideal = pd.read_excel(ideal_file, sheet_name=sheet)
                    df_transformed = pd.read_excel(transformed_file, sheet_name=sheet)
                    result, message = compare_two_dfs(df_ideal, df_transformed)

                    if not result:
                        return (False, f"Color Coding: {sheet_color_message}\nError: {message} in sheet={sheet}")
                return True, f"Color Coding: {sheet_color_message}"
            return (False, f"Sheets in the Excel are not same\t {wb1.sheetnames}!={wb2.sheetnames}")
        else:
            return (False, f"{ideal_file} not found in {ideal_folder}")
    except Exception as e:
        return False, f"Exception Occurred in File={transformed_file}={e} in sheet={cur_sheet}"

def regressionTest(ideal_folder, input_folder):
    comparison_files = [file for file in os.listdir(input_folder) if file.endswith(('xlsx', 'XLSX')) and not file.startswith(('expected_', 'original_'))]
    total = len(comparison_files)
    total_correct = 0
    start_time = time()

    with open(f"{input_folder}/regression_report.txt", "a") as f:
        for file in comparison_files:
            try:
                res, msg = compare_excel_files(f"{input_folder}/{file}", ideal_folder)
                f.write(f"{os.path.basename(file)}\n{msg}\n\n")

                total_correct += res
            except Exception as e:
                f.write(e)
        
        f.write(f"Test results for {input_folder}-\n")
        f.write(f"\n\nTotalFiles\tPassed\tFailed\n")
        f.write(f"{total}\t\t{total_correct}\t{total-total_correct}")
        f.write(f"\n\nTotal time taken={time()-start_time} seconds\n")
    
    return (total_correct == total)

#input_folder, ideal_folder = sys.argv[1], sys.argv[2]
#print(regressionTest(ideal_folder, input_folder))
