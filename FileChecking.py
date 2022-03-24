from threading import Thread
import openpyxl
import re
import pandas as pd
from TotalSheetChecking import miroTableChecking, totalsChecking

threadRes = {'scheme': '', 'exception': '', 'paypoint': '', 'miro': '', 'total': []}

# check for any scheme level errors
def schemeErrors(exceptionSheets, file):
    threadRes['scheme'] = "NONE"

    for sheet in exceptionSheets:
        excp_df = pd.read_excel(file, sheet_name=sheet)
        n = len(excp_df.index)
        for i in range(0, n):
            if(excp_df.iloc[i, 0].startswith('scheme')):
                threadRes['scheme'] = f"{excp_df.iloc[i, 1]} for {file}"

# check if mandatory columns exist 
def checkExceptionSheet(mcol, sheet, file):
    excpdf = pd.read_excel(file, sheet_name=sheet)
    n = len(excpdf.index)
    for i in range(0, n):
        if(mcol in excpdf.iloc[i, 1]):
            threadRes['exception'] = "NONE"

    threadRes['exception'] = f"{mcol} error not in exception sheet={sheet} which is absent in transformed sheet!"

# check for any exception sheet errors
def exceptionSheetError(transformedSheets, exceptionSheets, file):
    threadRes['exception'] = "NONE"

    for sheet in transformedSheets:
        transdf = pd.read_excel(file, sheet_name=sheet)

        mandatory_cols = {'unique ID': ['REFNO', 'PPSNO', 'PAYROLL', 'PPS NO'], 'Contribution': [
            'EE', 'ER', 'AVC', 'SPEE', 'SPER', 'SPAVC'], 'Surname': 'SURNAME', 'Forename': 'FORENAME'}
        trans_cols = transdf.keys()

        for mcol in mandatory_cols:
            res = any(col.upper() for col in trans_cols if col in mandatory_cols[mcol])
            if(len(transformedSheets) == 1):
                n_exp = len(exceptionSheets)
                if(not res and n_exp == 0):
                    threadRes['exception'] = f"Exception sheet not present for {sheet} in {file}! Mandatory column {mcol} not present."
                elif(not res and n_exp > 0):
                    checkExceptionSheet(mcol, exceptionSheets[0], file)
            else:
                if(not res and not f"${sheet}" in exceptionSheets):
                    threadRes['exception'] = f"Exception sheet not present for {sheet} in {file}! Mandatory column {mcol} not present."
                elif(not res and f"${sheet}" in exceptionSheets):
                    checkExceptionSheet(mcol, f"${sheet}", file)

# check for any paypoint errors
def paypointChecking(transformedSheets, file):
    threadRes['paypoint'] = "NONE"

    for sheet in transformedSheets:
        transdf = pd.read_excel(file, sheet_name=sheet)
        if('ISSUES FOUND' in transdf.columns and transdf['ISSUES FOUND'].isna().all()):
            threadRes['paypoint'] = f"Scheme Number not provided - Issues Found column is blank"
        elif('Paypoint' in transdf.columns):
            n = len(transdf.index)
            for i in range(0, n):
                if(pd.isna(transdf.iloc[i, -1]) and not ("paypoint is not returned" in transdf.iloc[i, 0] or "Member does not have open Paypoints" in transdf.iloc[i, 0])):
                    threadRes['paypoint'] = f"Paypoint in blank but no paypoint issue in ISSUES FOUND column in row {i} in sheet: {sheet}"
        else:
            if('ISSUES FOUND' not in transdf.columns):
                threadRes['paypoint'] = f"ISSUE FOUND column not present in sheet: {sheet}"
            else:
                n = len(transdf.index)
                for i in range(0, n):
                    if("Member does not have open Paypoints" not in transdf.iloc[i, 0]):
                        threadRes['paypoint'] = f"Scheme Number incorrect; Paypoint issue not present in ISSUE FOUND column in sheet: {sheet}"

# check for errors in transformed files
def errorChecking(filePath, parentRes):
    print("1. Error checking")
    res = ""
    wb = openpyxl.load_workbook(filePath)
    pattern = 'sheet "(.*?)">'
    allSheets = [re.search(pattern, str(ws)).group(1) for ws in wb.worksheets if ws.sheet_state != 'hidden']
    transformedSheets = [sheet for sheet in allSheets if not sheet.startswith(('$', '#', '~'))]
    exceptionSheets = [sheet for sheet in wb.sheetnames if sheet.startswith('$')]
    totalSheets = [sheet for sheet in wb.sheetnames if sheet.startswith('#')]

    thread_scheme = Thread(target=schemeErrors, args=(exceptionSheets, filePath), name='scheme')
    thread_paypoint = Thread(target=paypointChecking, args=(transformedSheets, filePath), name='paypoint')
    thread_exception = Thread(target=exceptionSheetError, args=(transformedSheets, exceptionSheets, filePath), name='exception')
    threads = [thread_scheme, thread_paypoint, thread_exception]
    if(totalSheets != []):
        thread_miro = Thread(target=miroTableChecking, args=(filePath, totalSheets, transformedSheets, threadRes), name='miro')
        threads.append(thread_miro)
        thread_total = Thread(target=totalsChecking, args=(filePath, totalSheets, threadRes), name='total')
        threads.append(thread_total)

    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

    res += f"Scheme Error: {threadRes['scheme']}\nException Sheet Error: {threadRes['exception']}\nPaypoint Error: {threadRes['paypoint']}\n"
    if(totalSheets != []):
        res += f"Miro Table Error: {threadRes['miro']}\n"
        if(threadRes['total'] != []):
            for msg in threadRes['total']:
                res += f"Totaling Check for {msg[0]}: {msg[1]}, {msg[2]}\n"

    parentRes['error'].append(res)